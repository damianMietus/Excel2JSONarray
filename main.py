import Tkinter
import tkMessageBox
from Tkinter import *
import Tkinter, Tkconstants, tkFileDialog
import sys
import string
import time

root = Tk()

# Parse Function
def parseFunc(directString, outputString, arrayString, column, mode, perline):

	try:
		from openpyxl import load_workbook
		from openpyxl import Workbook
	except:
		tkMessageBox.showinfo("Error", "Could not import openpyxl.")
		time.sleep(2)
		sys.exit(1)
		
	# Open Workbook
	try:
		wb = load_workbook(directString, data_only = True)
		ws = wb.active
		print ("Successfully opened the workbook")
	except:
		tkMessageBox.showinfo("Error", "Unable to open the sheet.")
		return

	index = "A1"
	itemList = []

	# Collect values from columns
	for i in range (1, ws.max_row+1):
		index = column+ str(i)
		if ws[index].value is None:
			endFlag = True
			break

		cellValue = ws[index].value

		# Apply text mode
		if (mode == "Lowercase"):
			cellValue = cellValue.lower()
		elif (mode == "Uppercase"):
			cellValue = cellValue.upper()

		print cellValue
		itemList.append(cellValue)

	exportStr = "{\n"+'"'+arrayString+'":['
	
	# Put list items into exportStr
	j = 0
	for i in range (0, len(itemList)):
	
		add = itemList[i]
		exportStr = exportStr+' "'+add+'"'

		if (i != len(itemList)-1):
			exportStr = exportStr+','

		j = j+1
		if (j == int(perline)):
			if (i != int(len(itemList))-1 ):
				exportStr = exportStr+'\n'
				j = 0

	exportStr = exportStr+"]\n}"

	f = open(outputString+".json", 'w')
	f.write(exportStr)
	f.close()

	tkMessageBox.showinfo("Success", "File successfully converted.")

#Button commands
def directoryFunc():
	root.filename = tkFileDialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))
	dirEntry.delete(0,END)
	dirEntry.insert(0, root.filename)

def checkVar():
	
	errorStr = "Failed to execute for the following reasons:\n"
	errorFlag = False

	dirvar = dirv.get()

	# Directory error
	if not dirvar:
		errorStr= errorStr+"- No directory selected\n"
		errorFlag = True

	outpvar = outpv.get()

	# Name.json error
	if not outpvar:
		errorStr= errorStr+"- No output file name given\n"
		errorFlag = True

	arvar = arv.get()

	# Array name error
	if not arvar:
		errorStr= errorStr+"- No json array name given\n"
		errorFlag = True

	exvar = exv.get()

	mode = op1Select.get()

	# Column error
	if not exvar:
		errorStr= errorStr+"- No column given\n"
		errorFlag = True
	elif (exvar.isalpha() == False):
		errorStr= errorStr+"- Column may only be alphabetical\n"
		errorFlag = True	

	perlinevar = perlinev.get()
	# Per Line Error
	if not perlinevar:
		perlinevar = "-1"
	elif (perlinevar.isdigit() == False):
		errorStr= errorStr+"- Items per line must be numerical\n"		
		errorFlag = True

	if (errorFlag == True):
		tkMessageBox.showinfo("Error", errorStr)
	else:
		# Success. Begin parsing
		parseFunc(dirvar, outpvar, arvar, exvar, mode, perlinevar)

root.geometry("500x175")
root.resizable(0, 0)
root.title("Excel2JSONarray")
root.pack_propagate(0)

rightFrame = Frame(root, width = 145, height = 165, bd = 2, relief=SUNKEN)
rightFrame.pack_propagate(0)
rightFrame.pack(side=RIGHT, anchor="n")

# Directory
dirLabel = Label(root, text="Choose file location:")
dirLabel.grid(row=0, column=0)

dirv = StringVar()
dirEntry = Entry(root, width=50,textvariable=dirv)
dirEntry.grid(row=3, column=0)

dirButton = Button(root, text="...", command=directoryFunc)
dirButton.grid(row=3, column = 1)

# Output name
outpLabel = Label(root, text="Output file name:")
outpLabel.grid(row=4, column = 0)

outpv = StringVar()
outpEntry = Entry(root, width=50, textvariable=outpv)
outpEntry.grid(row=5, column=0)

outpLabel = Label(root, text=".json")
outpLabel.grid(row=5, column = 1)

# Array name
arLabel = Label(root, text="Name of JSON array:")
arLabel.grid(row=6, column = 0)

arv = StringVar()
arEntry = Entry(root, width=50, textvariable=arv)
arEntry.grid(row=7, column=0)

# Excel Column
exLabel = Label(root, text="Excel column to read from:")
exLabel.grid(row=9, column = 0)

exv = StringVar()
exEntry = Entry(root, width=50, textvariable=exv)
exEntry.grid(row=10, column=0)

# Utilities frame

# Text Format
utilityLabel = Label(rightFrame, text="Text Format:",anchor="n")
utilityLabel.pack()

op1Vals = ["Default", "Lowercase", "Uppercase"]
op1Select = StringVar()
op1Select.set(op1Vals[0])

op1 = OptionMenu(rightFrame, op1Select, *op1Vals)
op1.pack()

# Tabs Per Line
perlineLabel = Label(rightFrame, text="Array Items Per Line:")
perlineLabel.pack()

perlinev = StringVar()
arEntry = Entry(rightFrame, width=5, textvariable=perlinev)
arEntry.pack()

# The Execute Button

excLabel = Label(rightFrame, text="")
excLabel.pack()
executeButton = Button(rightFrame, text="Execute", bd = 3, width=10, height=4, command = checkVar)
executeButton.pack(anchor="s")

root.mainloop()
